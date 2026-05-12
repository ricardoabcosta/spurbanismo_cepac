"""
Carga OUCAB → tabelas proposta, certidao, titulo_cepac, movimentacao.

Planilha : docs/novos/OUCAB_CONTROLE_ESTOQUES_mar_2026.xlsx
           Abas: A3, B, C, E1, E2, F1, F2, G, H, I1, I2 (+ Geral ignorada)
Banco    : DATABASE_URL em .env (psycopg2, síncrono)
OUC      : OUCAB (operacao_urbana_id = 3), Lei 15.893/2013

Layout das abas (colunas 0-indexadas):
  0  SETOR
  1  Nº DA PROPOSTA     (codigo AB-XXXX)
  2  Nº DA CERTIDÃO
  3  Nº PROCESSO SEI
  4  INTERESSADO
  5  Nº DOS CONTRIBUINTES
  6  R Inc  - Consumida   (R Incentivado / Unid. Habitacional Incentivada)
  7  R Inc  - Em analise
  8  R nInc - Consumida   (R Não-Incentivado / residencial padrão)
  9  R nInc - Em analise
  10 NR     - Consumida
  11 NR     - Em analise
  12 TOTAL  - Consumida
  13 TOTAL  - Em analise
  14 CEPAC-R  Cancelados
  15 CEPAC-R  Bloqueados
  16 CEPAC-nR Cancelados
  17 CEPAC-nR Bloqueados
  [18 CEPAC TOTAL Cancelados — só aba H]
  [19 CEPAC TOTAL Bloqueados — só aba H]
  18/20  STATUS DO PROCESSO
  19/21  Responsavel

Comportamento (três fases):

  FASE 1 — PROPOSTA (upsert por codigo)
    Proposta representativa = certidão com melhor status (DEFERIDO > ANALISE > CANCELADA),
    desempate pela data de emissão mais recente.
    Campos OUCAB: aca_r_inc_m2, aca_r_nao_inc_m2 (migration 032).

  FASE 2 — CERTIDÃO (upsert por numero_certidao)
    Todas as certidões encontradas nas abas. Situação:
      STATUS=Deferido  → 'VALIDA'
      STATUS=Cancelada → 'CANCELADA'
      STATUS=Analise   → 'ANALISE'

  FASE 3 — TituloCepac + Movimentacao (idempotente por codigo)
    Apenas certidões DEFERIDO com área > 0.
    R Não-Inc: uso=R, incentivado=False
    R Inc:     uso=R, incentivado=True
    NR:        uso=NR, incentivado=False

Validações finais (9 indicadores):
  1. Propostas ativas (DEFERIDO) = 6
  2. Certidões totais = 7
  3. Certidões VALIDA = 6
  4. TituloCepac CONSUMIDO = 6
  5. R Não-Inc consumido global = 152.985,48 m²   (≤ 1 m² divergência)
  6. NR consumido global = 4.380,03 m²             (≤ 1 m² divergência)
  7. Setor B R consumido = 152.985,48 m²
  8. Setor H NR consumido = 4.380,03 m²
  9. R Inc global = 0 m²

Uso:
    # Visualiza sem gravar:
    python scripts/carga_oucab.py --dry-run

    # Carga real (local):
    python scripts/carga_oucab.py

    # Produção (Azure):
    DATABASE_URL="postgresql+asyncpg://cepacadmin:SENHA@cepac-pgdb.postgres.database.azure.com:5432/cepac?ssl=require" \\
        python scripts/carga_oucab.py
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

try:
    import openpyxl
except ImportError:
    print("ERRO: instale openpyxl →  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERRO: instale psycopg2 →  pip install psycopg2-binary", file=sys.stderr)
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("carga_oucab")

# ---------------------------------------------------------------------------
# Caminhos
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent
XLSX_PATH = REPO_ROOT / "docs" / "novos" / "OUCAB_CONTROLE_ESTOQUES_mar_2026.xlsx"
ENV_PATH = REPO_ROOT / ".env"

# Abas com dados de propostas (ignora "Geral")
ABAS_SETOR = ["A3", "B", "C", "E1", "E2", "F1", "F2", "G", "H", "I1", "I2"]

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
OPERACAO_URBANA_ID = 3     # OUCAB
LEI_VIGENTE = "15.893/2013"
MOTIVO_CARGA = "CARGA_OUCAB_15893"
OPERADOR_SCRIPT = "script:carga_oucab"

# Mapeamento abreviatura → nome completo no banco
SETOR_NOME_MAP: dict[str, str] = {
    "A":  "Setor A",  "A1": "Setor A1", "A2": "Setor A2", "A3": "Setor A3",
    "B":  "Setor B",  "C":  "Setor C",  "D":  "Setor D",
    "E":  "Setor E",  "E1": "Setor E1", "E2": "Setor E2",
    "F":  "Setor F",  "F1": "Setor F1", "F2": "Setor F2",
    "G":  "Setor G",  "H":  "Setor H",
    "I":  "Setor I",  "I1": "Setor I1", "I2": "Setor I2",
}

STATUS_PA_MAP: dict[str, str] = {
    "DEFERIDO":   "DEFERIDO",
    "ANALISE":    "ANALISE",
    "EM ANALISE": "ANALISE",
    "CANCELADA":  "CANCELADA",
    "CANCELADO":  "CANCELADA",
    "INDEFERIDO": "INDEFERIDO",
}

SITUACAO_MAP: dict[str, str] = {
    "DEFERIDO":   "VALIDA",
    "ANALISE":    "ANALISE",
    "EM ANALISE": "ANALISE",
    "CANCELADA":  "CANCELADA",
    "CANCELADO":  "CANCELADA",
    "INDEFERIDO": "CANCELADA",
}

SITUACAO_PRIORIDADE: dict[str, int] = {
    "VALIDA":    0,
    "ANALISE":   1,
    "CANCELADA": 2,
}


# ---------------------------------------------------------------------------
# Leitura do .env
# ---------------------------------------------------------------------------

def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    with env_path.open() as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


# ---------------------------------------------------------------------------
# Helpers de parsing
# ---------------------------------------------------------------------------

def or_none(value) -> str | None:
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None


def parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        d = round(float(value), 2)
        return Decimal(str(d)) if d != 0 else Decimal("0")
    v = str(value).strip()
    if not v or v == "0":
        return Decimal("0")
    if "," in v:
        v = v.replace(".", "").replace(",", ".")
    try:
        return Decimal(v).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def parse_integer(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if value != 0 else None
    if isinstance(value, float):
        return int(value) if value != 0 else None
    v = str(value).strip().replace(".", "").replace(",", "")
    try:
        n = int(v)
        return n if n != 0 else None
    except ValueError:
        return None


def parse_area(value) -> Decimal:
    """Converte área para Decimal; retorna 0 se inválido ou nulo."""
    d = parse_decimal(value)
    return d if d is not None else Decimal("0")


# ---------------------------------------------------------------------------
# DATABASE_URL
# ---------------------------------------------------------------------------

def get_connect_kwargs() -> dict:
    raw = os.environ.get("DATABASE_URL", "")
    if not raw:
        raise RuntimeError(
            "DATABASE_URL não encontrada. "
            "Defina em .env ou exporte a variável de ambiente."
        )
    dsn = re.sub(r"^postgresql\+asyncpg://", "postgresql://", raw)
    connect_kwargs: dict = {}
    if "ssl=require" in dsn or "sslmode=require" in dsn:
        dsn = re.sub(r"[?&]ssl=require", "", dsn)
        dsn = re.sub(r"[?&]sslmode=require", "", dsn)
        dsn = re.sub(r"\?$", "", dsn)
        connect_kwargs["sslmode"] = "require"
    from urllib.parse import urlparse, unquote
    parsed = urlparse(dsn)
    if parsed.hostname:
        connect_kwargs.update({
            "host": parsed.hostname,
            "port": parsed.port or 5432,
            "dbname": parsed.path.lstrip("/"),
            "user": unquote(parsed.username or ""),
            "password": unquote(parsed.password or ""),
        })
    return connect_kwargs


# ---------------------------------------------------------------------------
# Estrutura de dados de uma linha da planilha
# ---------------------------------------------------------------------------

class LinhaOucab:
    """Linha parseada de uma aba da planilha OUCAB."""

    def __init__(self, row: tuple, status_col: int, aba: str) -> None:
        self.aba = aba
        self.setor_abrev: str = str(row[0] or "").strip()
        self.codigo: str = str(row[1] or "").strip()
        self.certidao: str = str(row[2] or "").strip()
        self.sei: str = str(row[3] or "").strip()
        self.interessado: str | None = or_none(row[4])
        self.contribuintes: str | None = or_none(row[5])

        self.r_inc_consumida: Decimal = parse_area(row[6])
        self.r_inc_analise:   Decimal = parse_area(row[7])
        self.r_ninc_consumida: Decimal = parse_area(row[8])
        self.r_ninc_analise:   Decimal = parse_area(row[9])
        self.nr_consumida: Decimal = parse_area(row[10])
        self.nr_analise:   Decimal = parse_area(row[11])
        self.total_consumida: Decimal = parse_area(row[12])
        self.total_analise:   Decimal = parse_area(row[13])

        cepac_r  = parse_integer(row[14]) or 0
        cepac_nr = parse_integer(row[16]) or 0
        self.cepac_aca: int | None = (cepac_r + cepac_nr) if (cepac_r + cepac_nr) > 0 else None

        raw_status = str(row[status_col] or "").strip().upper() if status_col < len(row) else ""
        self.status_raw: str = raw_status

        # Status normalizado
        self.status_pa: str = STATUS_PA_MAP.get(raw_status, "ANALISE")
        self.situacao: str = SITUACAO_MAP.get(raw_status, "ANALISE")

        # Setor nome completo
        self.setor_nome: str = SETOR_NOME_MAP.get(self.setor_abrev, f"Setor {self.setor_abrev}")

        # Derivados para proposta
        self.aca_r_inc_m2: Decimal | None = self.r_inc_consumida if self.r_inc_consumida > 0 else None
        self.aca_r_nao_inc_m2: Decimal | None = self.r_ninc_consumida if self.r_ninc_consumida > 0 else None
        self.aca_r_m2: Decimal | None = None
        r_total = self.r_inc_consumida + self.r_ninc_consumida
        if r_total > 0:
            self.aca_r_m2 = r_total
        self.aca_nr_m2: Decimal | None = self.nr_consumida if self.nr_consumida > 0 else None
        self.aca_total_m2: Decimal | None = self.total_consumida if self.total_consumida > 0 else None

        if r_total > 0 and self.nr_consumida > 0:
            self.uso_aca = "MISTO"
        elif r_total > 0:
            self.uso_aca = "R"
        elif self.nr_consumida > 0:
            self.uso_aca = "NR"
        else:
            self.uso_aca = None

    def is_valid_data_row(self) -> bool:
        """Verifica se a linha é um registro real (não template, subtotal, etc.)."""
        c = self.codigo
        if not c or c in ("XX-0000",) or c.startswith("modelo"):
            return False
        # Ignora linhas onde setor não é uma letra maiúscula válida
        if self.setor_abrev.lower() in ("subtotal", "total", "modelo/", "modelo"):
            return False
        if not re.match(r"^AB-\d{4}$", c):
            return False
        return True

    def __repr__(self) -> str:
        return (
            f"<LinhaOucab {self.codigo}/{self.certidao} "
            f"setor={self.setor_abrev} status={self.status_raw}>"
        )


# ---------------------------------------------------------------------------
# Seleção de linha representativa por proposta
# ---------------------------------------------------------------------------

def selecionar_representativa(linhas: list[LinhaOucab]) -> LinhaOucab:
    """Escolhe a linha representativa para preencher a proposta."""
    return min(
        linhas,
        key=lambda l: (
            SITUACAO_PRIORIDADE.get(l.situacao, 99),
            # Desempate: certidão mais recente (maior número léxico)
            -(int(re.search(r"\d+", l.certidao.split("/")[-1]).group())
              if re.search(r"\d+", l.certidao) else 0),
        ),
    )


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------

def ler_planilha(xlsx_path: Path) -> list[LinhaOucab]:
    """Lê todas as abas de setor e retorna lista de linhas válidas."""
    log.info("Lendo planilha: %s", xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    linhas: list[LinhaOucab] = []

    for aba in ABAS_SETOR:
        if aba not in wb.sheetnames:
            log.warning("Aba '%s' não encontrada — pulando.", aba)
            continue
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Detecta coluna STATUS (row index 2 = linha 3)
        header_row3 = rows[2]
        status_col = 18  # default (aba B, 21 cols)
        for i, v in enumerate(header_row3):
            if v is not None and "STATUS" in str(v).upper():
                status_col = i
                break

        # Linhas de dados a partir da linha 5 (index 4)
        for row in rows[4:]:
            if not row or all(v is None for v in row):
                continue
            # Precisa ter pelo menos colunas até status_col
            if len(row) <= 1:
                continue
            linha = LinhaOucab(row, status_col, aba)
            # Filtra apenas linhas da aba correspondente (evita duplicatas cross-aba)
            if linha.setor_abrev != aba:
                continue
            if not linha.is_valid_data_row():
                continue
            linhas.append(linha)
            log.debug("  Lida: %s", linha)

    wb.close()
    log.info("Total de linhas válidas: %d", len(linhas))
    return linhas


# ---------------------------------------------------------------------------
# SQLs de upsert
# ---------------------------------------------------------------------------

UPSERT_PROPOSTA = """
    INSERT INTO proposta (
        codigo, numero_pa, tipo_processo, status_pa,
        interessado, setor_id, requerimento,
        uso_aca, aca_r_m2, aca_nr_m2, aca_total_m2,
        aca_r_inc_m2, aca_r_nao_inc_m2,
        tipo_contrapartida, cepac_aca, cepac_total,
        certidao, situacao_certidao, data_certidao,
        lei_vigente
    ) VALUES (
        %(codigo)s, %(numero_pa)s, %(tipo_processo)s, %(status_pa)s,
        %(interessado)s, %(setor_id)s::uuid, %(requerimento)s,
        %(uso_aca)s, %(aca_r_m2)s, %(aca_nr_m2)s, %(aca_total_m2)s,
        %(aca_r_inc_m2)s, %(aca_r_nao_inc_m2)s,
        %(tipo_contrapartida)s, %(cepac_aca)s, %(cepac_total)s,
        %(certidao)s, %(situacao_certidao)s, %(data_certidao)s,
        %(lei_vigente)s
    )
    ON CONFLICT (codigo) DO UPDATE SET
        numero_pa           = EXCLUDED.numero_pa,
        tipo_processo       = EXCLUDED.tipo_processo,
        status_pa           = EXCLUDED.status_pa,
        interessado         = EXCLUDED.interessado,
        setor_id            = EXCLUDED.setor_id,
        uso_aca             = EXCLUDED.uso_aca,
        aca_r_m2            = EXCLUDED.aca_r_m2,
        aca_nr_m2           = EXCLUDED.aca_nr_m2,
        aca_total_m2        = EXCLUDED.aca_total_m2,
        aca_r_inc_m2        = EXCLUDED.aca_r_inc_m2,
        aca_r_nao_inc_m2    = EXCLUDED.aca_r_nao_inc_m2,
        tipo_contrapartida  = EXCLUDED.tipo_contrapartida,
        cepac_aca           = EXCLUDED.cepac_aca,
        cepac_total         = EXCLUDED.cepac_total,
        certidao            = EXCLUDED.certidao,
        situacao_certidao   = EXCLUDED.situacao_certidao,
        data_certidao       = EXCLUDED.data_certidao,
        lei_vigente         = EXCLUDED.lei_vigente,
        updated_at          = now()
    RETURNING id, xmax
"""

UPSERT_CERTIDAO = """
    INSERT INTO certidao (
        proposta_id, numero_certidao, tipo, situacao,
        numero_processo_sei, uso_aca,
        aca_r_m2, aca_nr_m2, aca_total_m2,
        tipo_contrapartida, cepac_aca, cepac_total, obs
    ) VALUES (
        %(proposta_id)s::uuid, %(numero_certidao)s,
        %(tipo)s::tipo_certidao_enum, %(situacao)s::situacao_certidao_enum,
        %(numero_processo_sei)s, %(uso_aca)s,
        %(aca_r_m2)s, %(aca_nr_m2)s, %(aca_total_m2)s,
        %(tipo_contrapartida)s, %(cepac_aca)s, %(cepac_total)s, %(obs)s
    )
    ON CONFLICT (numero_certidao) DO UPDATE SET
        proposta_id         = EXCLUDED.proposta_id,
        tipo                = EXCLUDED.tipo,
        situacao            = EXCLUDED.situacao,
        numero_processo_sei = EXCLUDED.numero_processo_sei,
        uso_aca             = EXCLUDED.uso_aca,
        aca_r_m2            = EXCLUDED.aca_r_m2,
        aca_nr_m2           = EXCLUDED.aca_nr_m2,
        aca_total_m2        = EXCLUDED.aca_total_m2,
        tipo_contrapartida  = EXCLUDED.tipo_contrapartida,
        cepac_aca           = EXCLUDED.cepac_aca,
        cepac_total         = EXCLUDED.cepac_total,
        obs                 = EXCLUDED.obs
    WHERE certidao.situacao != 'VALIDA'::situacao_certidao_enum
       OR EXCLUDED.situacao = 'VALIDA'::situacao_certidao_enum
    RETURNING xmax
"""

UPSERT_TITULO = """
    INSERT INTO titulo_cepac (id, codigo, setor_id, valor_m2, uso, origem, estado, incentivado)
    VALUES (%(id)s::uuid, %(codigo)s, %(setor_id)s::uuid, %(valor_m2)s,
            %(uso)s::uso_enum, 'ACA'::origem_enum, %(estado)s::estado_titulo_enum, %(incentivado)s)
    ON CONFLICT (codigo) DO NOTHING
    RETURNING id
"""

INSERT_MOVIMENTACAO = """
    INSERT INTO movimentacao (
        titulo_id, setor_id, uso, origem,
        estado_anterior, estado_novo,
        numero_processo_sei, motivo, incentivado, operador
    ) VALUES (
        %(titulo_id)s::uuid, %(setor_id)s::uuid,
        %(uso)s::uso_enum, 'ACA'::origem_enum,
        NULL, %(estado_novo)s::estado_titulo_enum,
        %(numero_processo_sei)s, %(motivo)s, %(incentivado)s, %(operador)s
    )
"""


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def carregar(connect_kwargs: dict, xlsx_path: Path, dry_run: bool) -> None:
    linhas = ler_planilha(xlsx_path)
    if not linhas:
        log.error("Nenhuma linha válida encontrada na planilha.")
        sys.exit(1)

    # Agrupa por codigo
    grupos: dict[str, list[LinhaOucab]] = defaultdict(list)
    for l in linhas:
        grupos[l.codigo].append(l)

    log.info("Propostas únicas: %d  |  Certidões: %d", len(grupos), len(linhas))

    # Conecta ao banco
    log.info("Conectando ao banco (dry_run=%s)...", dry_run)
    conn = psycopg2.connect(**connect_kwargs)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Pre-load: setores OUCAB → {nome_completo: id}
    cur.execute(
        "SELECT id, nome FROM setor WHERE operacao_urbana_id = %s",
        (OPERACAO_URBANA_ID,),
    )
    setor_map: dict[str, str] = {r["nome"]: str(r["id"]) for r in cur.fetchall()}
    log.info("Setores OUCAB carregados: %s", sorted(setor_map.keys()))

    if not setor_map:
        log.error("Nenhum setor OUCAB encontrado (operacao_urbana_id=%d).", OPERACAO_URBANA_ID)
        conn.close()
        sys.exit(1)

    # -----------------------------------------------------------------------
    # FASE 1 — PROPOSTA
    # -----------------------------------------------------------------------
    cnt_prop_insert = cnt_prop_update = cnt_prop_skip = 0
    proposta_id_map: dict[str, str] = {}   # codigo → UUID

    for codigo, grupo in sorted(grupos.items()):
        rep = selecionar_representativa(grupo)

        setor_id = setor_map.get(rep.setor_nome)
        if not setor_id:
            log.warning("SKIP proposta %s: setor '%s' não encontrado.", codigo, rep.setor_nome)
            cnt_prop_skip += 1
            continue

        params = {
            "codigo": codigo,
            "numero_pa": rep.sei or None,
            "tipo_processo": "SEI" if rep.sei else None,
            "status_pa": rep.status_pa,
            "interessado": rep.interessado,
            "setor_id": setor_id,
            "requerimento": "VINCULACAO",
            "uso_aca": rep.uso_aca,
            "aca_r_m2": str(rep.aca_r_m2) if rep.aca_r_m2 else None,
            "aca_nr_m2": str(rep.aca_nr_m2) if rep.aca_nr_m2 else None,
            "aca_total_m2": str(rep.aca_total_m2) if rep.aca_total_m2 else None,
            "aca_r_inc_m2": str(rep.aca_r_inc_m2) if rep.aca_r_inc_m2 else None,
            "aca_r_nao_inc_m2": str(rep.aca_r_nao_inc_m2) if rep.aca_r_nao_inc_m2 else None,
            "tipo_contrapartida": "CEPAC (título)",
            "cepac_aca": rep.cepac_aca,
            "cepac_total": rep.cepac_aca,
            "certidao": rep.certidao[:20] if rep.certidao else None,
            "situacao_certidao": rep.situacao,
            "data_certidao": None,  # OUCAB planilha não tem data de emissão
            "lei_vigente": LEI_VIGENTE,
        }

        if dry_run:
            proposta_id_map[codigo] = str(uuid4())
            cnt_prop_insert += 1
            log.info("  [DRY] PROPOSTA %s  setor=%s  status=%s  R=%.2f  NR=%.2f  CEPACs=%s",
                     codigo, rep.setor_nome, rep.status_pa,
                     float(rep.aca_r_m2 or 0), float(rep.aca_nr_m2 or 0), rep.cepac_aca)
            continue

        try:
            cur.execute(UPSERT_PROPOSTA, params)
            result = cur.fetchone()
            proposta_id_map[codigo] = str(result["id"])
            if result["xmax"] == 0:
                cnt_prop_insert += 1
            else:
                cnt_prop_update += 1
        except Exception as exc:
            conn.rollback()
            log.error("ERRO proposta %s: %s", codigo, exc)
            cnt_prop_skip += 1

    if not dry_run:
        conn.commit()

    log.info("Fase 1: %d inseridas, %d atualizadas, %d puladas",
             cnt_prop_insert, cnt_prop_update, cnt_prop_skip)

    # -----------------------------------------------------------------------
    # FASE 2 — CERTIDÃO
    # -----------------------------------------------------------------------
    cnt_cert_insert = cnt_cert_update = cnt_cert_skip = 0

    for linha in sorted(linhas, key=lambda l: (
        SITUACAO_PRIORIDADE.get(l.situacao, 99),  # CANCELADA primeiro, VALIDA por último
    ), reverse=True):
        proposta_id = proposta_id_map.get(linha.codigo)
        if not proposta_id:
            cnt_cert_skip += 1
            continue

        numero_certidao = linha.certidao[:20] if linha.certidao else f"AB-{linha.codigo}-sem"

        params = {
            "proposta_id": proposta_id,
            "numero_certidao": numero_certidao,
            "tipo": "VINCULAÇÃO",
            "situacao": linha.situacao,
            "numero_processo_sei": linha.sei or None,
            "uso_aca": linha.uso_aca,
            "aca_r_m2": str(linha.aca_r_m2) if linha.aca_r_m2 else None,
            "aca_nr_m2": str(linha.aca_nr_m2) if linha.aca_nr_m2 else None,
            "aca_total_m2": str(linha.aca_total_m2) if linha.aca_total_m2 else None,
            "tipo_contrapartida": "CEPAC (título)",
            "cepac_aca": linha.cepac_aca,
            "cepac_total": linha.cepac_aca,
            "obs": f"Responsável: {or_none('Planilha OUCAB mar/2026')}",
        }

        if dry_run:
            cnt_cert_insert += 1
            log.info("  [DRY] CERTIDÃO %s  (%s)  situacao=%s  R=%.2f  NR=%.2f",
                     numero_certidao, linha.codigo, linha.situacao,
                     float(linha.aca_r_m2 or 0), float(linha.aca_nr_m2 or 0))
            continue

        try:
            cur.execute(UPSERT_CERTIDAO, params)
            result = cur.fetchone()
            if result is None:
                cnt_cert_skip += 1     # WHERE condition not met
            elif result["xmax"] == 0:
                cnt_cert_insert += 1
            else:
                cnt_cert_update += 1
        except Exception as exc:
            conn.rollback()
            log.error("ERRO certidão %s (%s): %s", numero_certidao, linha.codigo, exc)
            cnt_cert_skip += 1

    if not dry_run:
        conn.commit()

    log.info("Fase 2: %d inseridas, %d atualizadas, %d puladas",
             cnt_cert_insert, cnt_cert_update, cnt_cert_skip)

    # -----------------------------------------------------------------------
    # FASE 3 — TituloCepac + Movimentacao
    # -----------------------------------------------------------------------
    cnt_titulo_insert = cnt_titulo_skip = 0

    for linha in linhas:
        # Apenas certidões DEFERIDO (VALIDA) com área > 0
        if linha.situacao != "VALIDA":
            continue

        setor_id = setor_map.get(linha.setor_nome)
        if not setor_id:
            continue

        # Dimensões: R Inc, R Não-Inc, NR
        dimensoes: list[tuple[str, Decimal, bool]] = []  # (uso, area, incentivado)
        if linha.r_inc_consumida > 0:
            dimensoes.append(("R", linha.r_inc_consumida, True))
        if linha.r_ninc_consumida > 0:
            dimensoes.append(("R", linha.r_ninc_consumida, False))
        if linha.nr_consumida > 0:
            dimensoes.append(("NR", linha.nr_consumida, False))

        if not dimensoes:
            continue

        for uso, area, incentivado in dimensoes:
            inc_suffix = "Inc" if incentivado else "nInc"
            titulo_codigo = f"OUCAB-{linha.certidao}-{uso}-{inc_suffix}"

            if dry_run:
                cnt_titulo_insert += 1
                log.info(
                    "  [DRY] TITULO %s  uso=%s  incentivado=%s  area=%.2f  setor=%s",
                    titulo_codigo, uso, incentivado, float(area), linha.setor_nome,
                )
                continue

            titulo_id = str(uuid4())
            try:
                cur.execute(UPSERT_TITULO, {
                    "id": titulo_id,
                    "codigo": titulo_codigo[:50],
                    "setor_id": setor_id,
                    "valor_m2": str(area),
                    "uso": uso,
                    "estado": "CONSUMIDO",
                    "incentivado": incentivado,
                })
                result = cur.fetchone()
                if result is None:
                    cnt_titulo_skip += 1
                    continue  # ON CONFLICT DO NOTHING

                titulo_real_id = str(result["id"])
                cur.execute(INSERT_MOVIMENTACAO, {
                    "titulo_id": titulo_real_id,
                    "setor_id": setor_id,
                    "uso": uso,
                    "estado_novo": "CONSUMIDO",
                    "numero_processo_sei": (linha.sei or "CARGA_OUCAB")[:50],
                    "motivo": MOTIVO_CARGA,
                    "incentivado": incentivado,
                    "operador": OPERADOR_SCRIPT,
                })
                cnt_titulo_insert += 1
            except Exception as exc:
                conn.rollback()
                log.error("ERRO titulo %s: %s", titulo_codigo, exc)
                cnt_titulo_skip += 1

    if not dry_run:
        conn.commit()

    log.info("Fase 3: %d TituloCepac inseridos, %d já existiam",
             cnt_titulo_insert, cnt_titulo_skip)

    conn.close()

    # -----------------------------------------------------------------------
    # Validações e relatório final
    # -----------------------------------------------------------------------
    _relatorio(linhas, grupos, dry_run, cnt_prop_insert, cnt_prop_update, cnt_prop_skip,
               cnt_cert_insert, cnt_cert_update, cnt_cert_skip,
               cnt_titulo_insert, cnt_titulo_skip)


def _relatorio(
    linhas: list[LinhaOucab],
    grupos: dict[str, list[LinhaOucab]],
    dry_run: bool,
    cnt_prop_insert: int, cnt_prop_update: int, cnt_prop_skip: int,
    cnt_cert_insert: int, cnt_cert_update: int, cnt_cert_skip: int,
    cnt_titulo_insert: int, cnt_titulo_skip: int,
) -> None:
    # Cálculos de validação a partir dos dados da planilha
    r_ninc_total = sum(l.r_ninc_consumida for l in linhas if l.situacao == "VALIDA")
    r_inc_total  = sum(l.r_inc_consumida  for l in linhas if l.situacao == "VALIDA")
    nr_total     = sum(l.nr_consumida     for l in linhas if l.situacao == "VALIDA")

    r_setor: dict[str, Decimal] = defaultdict(Decimal)
    nr_setor: dict[str, Decimal] = defaultdict(Decimal)
    for l in linhas:
        if l.situacao != "VALIDA":
            continue
        r_setor[l.setor_nome] += l.r_inc_consumida + l.r_ninc_consumida
        nr_setor[l.setor_nome] += l.nr_consumida

    prop_deferidas = sum(1 for g in grupos.values()
                         if selecionar_representativa(g).status_pa == "DEFERIDO")
    cert_validas   = sum(1 for l in linhas if l.situacao == "VALIDA")

    print()
    print("=" * 70)
    print("  RESULTADO DA CARGA OUCAB" + ("  [DRY-RUN — nenhum dado gravado]" if dry_run else ""))
    print("=" * 70)
    print()
    print("  PROPOSTAS")
    print(f"    Inseridas                  : {cnt_prop_insert:>6}")
    print(f"    Atualizadas                : {cnt_prop_update:>6}")
    print(f"    Puladas/erro               : {cnt_prop_skip:>6}")
    print()
    print("  CERTIDÕES")
    print(f"    Inseridas                  : {cnt_cert_insert:>6}")
    print(f"    Atualizadas                : {cnt_cert_update:>6}")
    print(f"    Puladas                    : {cnt_cert_skip:>6}")
    print()
    print("  TITULO_CEPAC + MOVIMENTACAO")
    print(f"    Inseridos                  : {cnt_titulo_insert:>6}")
    print(f"    Já existiam (skip)         : {cnt_titulo_skip:>6}")
    print()
    print("  VALIDAÇÕES (9 indicadores — tolerância ≤ 1 m²):")
    print()

    V_PROP     = 6
    V_CERT     = 7
    V_CERT_V   = 6
    V_TITULO   = 6
    V_R_NINC   = Decimal("152985.48")
    V_NR       = Decimal("4380.03")
    V_R_B      = Decimal("152985.48")
    V_NR_H     = Decimal("4380.03")
    V_R_INC    = Decimal("0")

    checks = [
        ("1. Propostas DEFERIDO",         prop_deferidas,            V_PROP,   "== "),
        ("2. Certidões totais",            len(linhas),               V_CERT,   "== "),
        ("3. Certidões VALIDA",            cert_validas,              V_CERT_V, "== "),
        ("4. TituloCepac CONSUMIDO",       cnt_titulo_insert + cnt_titulo_skip, V_TITULO, "== "),
        ("5. R Não-Inc global (m²)",       float(r_ninc_total),       float(V_R_NINC), "~= "),
        ("6. NR global (m²)",              float(nr_total),           float(V_NR),     "~= "),
        ("7. Setor B R consumido (m²)",    float(r_setor.get("Setor B", Decimal("0"))), float(V_R_B), "~= "),
        ("8. Setor H NR consumido (m²)",   float(nr_setor.get("Setor H", Decimal("0"))), float(V_NR_H), "~= "),
        ("9. R Inc global (m²)",           float(r_inc_total),        float(V_R_INC), "== "),
    ]

    all_ok = True
    for label, got, expected, op in checks:
        if op == "== ":
            ok = got == expected
        else:  # ~=
            ok = abs(float(got) - float(expected)) <= 1.0
        mark = "✓" if ok else "✗"
        if not ok:
            all_ok = False
        if isinstance(got, float):
            print(f"  {mark} {label:<35} {got:>12.2f}  (esperado {expected:.2f})")
        else:
            print(f"  {mark} {label:<35} {got:>12}  (esperado {expected})")

    print()
    print("  TOTAIS POR SETOR (certidões VALIDA):")
    print("  {:<15} {:>15} {:>15}".format("SETOR", "R Consumido", "NR Consumido"))
    print("  " + "-" * 47)
    todos_setores = sorted(set(list(r_setor.keys()) + list(nr_setor.keys())))
    for s in todos_setores:
        print("  {:<15} {:>15.2f} {:>15.2f}".format(
            s, float(r_setor.get(s, Decimal("0"))), float(nr_setor.get(s, Decimal("0")))
        ))

    print()
    if all_ok:
        print("  STATUS: ✓ Todos os 9 indicadores aprovados.")
    else:
        print("  STATUS: ✗ Indicadores com divergência — revisar antes de produção.")
    print()
    print("=" * 70)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Carga OUCAB no banco CEPAC")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Calcula e exibe os totais sem gravar no banco",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=XLSX_PATH,
        help="Caminho alternativo para a planilha XLSX",
    )
    args = parser.parse_args()

    load_env(ENV_PATH)

    xlsx = args.xlsx
    if not xlsx.exists():
        log.error("Planilha não encontrada: %s", xlsx)
        sys.exit(1)

    if not args.dry_run:
        try:
            connect_kwargs = get_connect_kwargs()
        except RuntimeError as e:
            log.error(str(e))
            sys.exit(1)
        log.info("Banco: %s@%s/%s",
                 connect_kwargs.get("user"),
                 connect_kwargs.get("host"),
                 connect_kwargs.get("dbname"))
    else:
        # dry-run: não precisa de banco real, mas conectar_kwargs é necessário para _relatorio
        try:
            connect_kwargs = get_connect_kwargs()
        except RuntimeError:
            connect_kwargs = {}

    carregar(connect_kwargs, xlsx, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
