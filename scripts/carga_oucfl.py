"""
Carga da planilha OUCFL → tabelas `proposta`, `certidao`,
`titulo_cepac` e `movimentacao` (PostgreSQL).

Planilha : docs/novos/OUCFL_ESTOQUE_abr_rv02.xlsx  aba "2_CONTROLE_ESTOQUE"
Banco     : DATABASE_URL em .env (psycopg2, síncrono)

Comportamento (três fases):

  FASE 1 — PROPOSTA (uma linha por FL-XXXX único)
    Seleciona linha representativa por codigo:
      1. SITUACAO = VALIDA    → a mais recente por DATA CERTIDAO
      2. SITUACAO = ANALISE   → a mais recente por DATA CERTIDAO
      3. Senão (CANCELADA)    → a mais recente por DATA CERTIDAO
    UPSERT em proposta com todos os campos, incluindo campos novos
    da migration 028 (lei_vigente, aca_r_real_m2, aca_r_beneficios_m2,
    aca_nr_real_m2, aca_nr_beneficios_m2).
    Áreas ACA NET: usa coluna "ACA-R" quando preenchida; caso contrário
    computa REAL - BENEFICIOS.

  FASE 2 — CERTIDÃO (uma linha por linha da planilha)
    UPSERT em certidao ON CONFLICT (numero_certidao).
    Linhas sem certidão recebem fallback FL-{codigo}-{idx}.

  FASE 3A — TituloCepac + Movimentacao para Lei 13.769/2004 (histórico)
    Um único TituloCepac(CONSUMIDO) por setor por dimensão (R e NR) com o
    consumo líquido consolidado extraído da aba Consolidado_OUC-FL.
    Idempotência por titulo_cepac.codigo UNIQUE.

  FASE 3B — TituloCepac + Movimentacao para Lei 18.175/2024 (individual)
    Por certidão com VINCULACAO / ALTERACAO e área > 0:
      - DEFERIDO + VALIDA  → TituloCepac(CONSUMIDO) + Movimentacao
      - ANALISE            → TituloCepac(EM_ANALISE) + Movimentacao
    DESVINCULACAO é ignorada.
    Idempotência por titulo_cepac.codigo UNIQUE.

Uso:
    # Visualiza totais sem inserir no banco:
    python scripts/carga_oucfl.py --dry-run

    # Carga real (local):
    python scripts/carga_oucfl.py

    # Produção (Azure):
    DATABASE_URL="postgresql+asyncpg://cepacadmin:SENHA@cepac-pgdb.postgres.database.azure.com:5432/cepac?ssl=require" \\
        python scripts/carga_oucfl.py
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

try:
    import openpyxl
except ImportError:
    print("ERRO: instale openpyxl →  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERRO: instale psycopg2 →  pip install psycopg2-binary", file=sys.stderr)
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("carga_oucfl")

# ---------------------------------------------------------------------------
# Caminhos
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent
XLSX_PATH = REPO_ROOT / "docs" / "novos" / "OUCFL_ESTOQUE_abr_rv02.xlsx"
ENV_PATH = REPO_ROOT / ".env"

ABA = "2_CONTROLE_ESTOQUE"

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
LEI_VIGENTE = "18.175/2024"
OPERADOR_SCRIPT = "script:carga_oucfl"
MOTIVO_CARGA_18175 = "CARGA_OUCFL_18175"
MOTIVO_CARGA_13769 = "CARGA_OUCFL_13769"

# Consumo líquido histórico por setor sob a Lei 13.769/2004.
# Fonte: planilha Consolidado_OUC-FL, linha 12 (Total R/NR) → linha 8-11 por setor.
# Valores = vinculações - desvinculações sob a Lei 13.769 (já inclui FL-001/2015 etc.)
HISTORICO_13769: dict[str, dict[str, Decimal]] = {
    "HÉLIO PELEGRINO": {"R": Decimal("279084.28"), "NR": Decimal("162600.47")},
    "FARIA LIMA":      {"R": Decimal("133370.09"), "NR": Decimal("73552.78")},
    "PINHEIROS":       {"R": Decimal("146445.55"), "NR": Decimal("95217.97")},
    "OLIMPÍADAS":      {"R": Decimal("87199.88"),  "NR": Decimal("95563.02")},
}

STATUS_PA_MAP: dict[str, str] = {
    "ANALISE":    "ANALISE",
    "DEFERIDO":   "DEFERIDO",
    "INDEFERIDO": "INDEFERIDO",
}

REQUERIMENTO_MAP: dict[str, str] = {
    "VINCULACAO":    "VINCULACAO",
    "ALTERACAO":     "ALTERACAO",
    "DESVINCULACAO": "DESVINCULACAO",
}

TIPO_CERTIDAO_MAP: dict[str, str] = {
    "VINCULACAO":    "VINCULAÇÃO",
    "ALTERACAO":     "ALTERAÇÃO",
    "DESVINCULACAO": "DESVINCULAÇÃO",
}

SITUACAO_MAP: dict[str, str] = {
    "ANALISE":   "ANALISE",
    "VALIDA":    "VALIDA",
    "CANCELADA": "CANCELADA",
}

SITUACAO_PRIORIDADE: dict[str, int] = {
    "VALIDA":    0,
    "ANALISE":   1,
    "CANCELADA": 2,
}


# ---------------------------------------------------------------------------
# Leitura do .env
# ---------------------------------------------------------------------------

def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    with env_path.open() as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


# ---------------------------------------------------------------------------
# Helpers de parsing
# ---------------------------------------------------------------------------

def or_none(value) -> str | None:
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None


def clean_semicolon(value) -> str | None:
    if value is None:
        return None
    v = str(value).strip().rstrip(";").strip()
    return v if v else None


def parse_decimal(value) -> Decimal | None:
    """Converte int/float/str para Decimal. Aceita PT-BR e ponto decimal."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(round(value, 2)))
    v = str(value).strip()
    if not v:
        return None
    # Tratar formato PT-BR (ponto de milhar, vírgula decimal)
    if "," in v:
        v = v.replace(".", "").replace(",", ".")
    try:
        return Decimal(v)
    except InvalidOperation:
        return None


def parse_integer(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    v = str(value).strip().replace(".", "").replace(",", "")
    try:
        return int(v)
    except ValueError:
        return None


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    v = str(value).strip()
    if not v:
        return None
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", v)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def detect_tipo_processo(pa16, pa12) -> tuple[str | None, str | None]:
    sei = or_none(pa16)
    simproc = or_none(pa12)
    if sei:
        return sei, "SEI"
    if simproc:
        return simproc, "SIMPROC"
    return None, None


def compute_net_aca(net_col, real_col, ben_col) -> Decimal | None:
    """Retorna área ACA líquida: usa coluna net se disponível, senão REAL - BEN."""
    if net_col is not None:
        d = parse_decimal(net_col)
        if d is not None and d > 0:
            return d
    real = parse_decimal(real_col)
    if real is not None and real > 0:
        ben = parse_decimal(ben_col) or Decimal("0")
        net = real - ben
        return net if net > 0 else Decimal("0")
    return None


# ---------------------------------------------------------------------------
# DATABASE_URL
# ---------------------------------------------------------------------------

def get_connect_kwargs() -> dict:
    raw = os.environ.get("DATABASE_URL", "")
    if not raw:
        raise RuntimeError(
            "DATABASE_URL não encontrada. "
            "Defina em .env ou exporte a variável de ambiente."
        )
    dsn = re.sub(r"^postgresql\+asyncpg://", "postgresql://", raw)
    connect_kwargs: dict = {}
    if "ssl=require" in dsn or "sslmode=require" in dsn:
        dsn = re.sub(r"[?&]ssl=require", "", dsn)
        dsn = re.sub(r"[?&]sslmode=require", "", dsn)
        dsn = re.sub(r"\?$", "", dsn)
        connect_kwargs["sslmode"] = "require"
    from urllib.parse import urlparse, unquote
    parsed = urlparse(dsn)
    if parsed.hostname:
        connect_kwargs.update({
            "host": parsed.hostname,
            "port": parsed.port or 5432,
            "dbname": parsed.path.lstrip("/"),
            "user": unquote(parsed.username or ""),
            "password": unquote(parsed.password or ""),
        })
    return connect_kwargs


# ---------------------------------------------------------------------------
# Seleção de linha representativa por codigo
# ---------------------------------------------------------------------------

def selecionar_representativa(
    linhas: list[tuple[int, dict]],
) -> dict:
    melhor = max(
        linhas,
        key=lambda item: (
            -SITUACAO_PRIORIDADE.get(
                str(item[1].get("SITUACAO") or "").strip().upper(), 99
            ),
            item[1].get("_data_cert") or date.min,
            item[0],
        ),
    )
    return melhor[1]


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def carregar(connect_kwargs: dict, xlsx_path: Path, dry_run: bool) -> None:
    # Carregar planilha
    log.info("Lendo planilha: %s  aba: %s", xlsx_path, ABA)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        log.error("Aba '%s' não encontrada. Abas disponíveis: %s", ABA, wb.sheetnames)
        sys.exit(1)
    ws = wb[ABA]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        log.error("Planilha vazia.")
        sys.exit(1)

    header = all_rows[0]
    col: dict[str, int] = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def cell(row: tuple, col_name: str):
        idx = col.get(col_name)
        if idx is None:
            return None
        return row[idx] if idx < len(row) else None

    # Linhas com CODIGO preenchido
    raw_rows = [
        r for r in all_rows[1:]
        if r and r[col.get("CODIGO", 0)] is not None
        and str(r[col.get("CODIGO", 0)]).strip()
    ]
    log.info("Linhas com CODIGO preenchido: %d", len(raw_rows))

    # Enriquece cada linha com data parseada para facilitar seleção
    rows_enriched: list[dict] = []
    for raw in raw_rows:
        d: dict = {h: cell(raw, h) for h in col}
        d["_data_cert"] = parse_date(cell(raw, "DATA CERTIDAO "))
        rows_enriched.append(d)

    # -----------------------------------------------------------------------
    # Conectar ao banco
    # -----------------------------------------------------------------------
    log.info("Conectando ao banco (dry_run=%s)...", dry_run)
    conn = psycopg2.connect(**connect_kwargs)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # -----------------------------------------------------------------------
    # Pre-load: setores OUCFL (operacao_urbana_id = 2) → {UPPER(nome): id}
    # -----------------------------------------------------------------------
    cur.execute(
        "SELECT id, nome FROM setor WHERE operacao_urbana_id = 2 AND ativo = TRUE"
    )
    setor_map: dict[str, str] = {}
    for r in cur.fetchall():
        setor_map[r["nome"].upper()] = str(r["id"])
    log.info("Setores OUCFL: %s", list(setor_map.keys()))

    if not setor_map:
        log.error(
            "Nenhum setor ativo encontrado para OUCFL (operacao_urbana_id=2). "
            "Verifique se as migrations foram executadas."
        )
        conn.close()
        sys.exit(1)

    # -----------------------------------------------------------------------
    # Grupos por codigo para representativa
    # -----------------------------------------------------------------------
    grupos: dict[str, list[tuple[int, dict]]] = defaultdict(list)
    for idx, row in enumerate(rows_enriched):
        codigo = str(row.get("CODIGO") or "").strip()
        if codigo:
            grupos[codigo].append((idx, row))

    log.info("Códigos únicos: %d", len(grupos))

    # -----------------------------------------------------------------------
    # FASE 1 — PROPOSTA
    # -----------------------------------------------------------------------
    UPSERT_PROPOSTA = """
        INSERT INTO proposta (
            codigo, numero_pa, tipo_processo, data_autuacao, status_pa,
            interessado, cnpj, cpf, tipo_interessado, endereco,
            setor_id, requerimento,
            contribuinte_sq, contribuinte_lote, area_terreno_m2,
            uso_aca, aca_r_m2, aca_nr_m2, aca_total_m2,
            tipo_contrapartida, valor_oodc_rs,
            cepac_aca, cepac_parametros, cepac_total,
            certidao, situacao_certidao, data_certidao,
            nuvem_r_m2, nuvem_nr_m2, nuvem_total_m2, nuvem_cepac,
            obs, resp_data, data_proposta,
            lei_vigente,
            aca_r_real_m2, aca_r_beneficios_m2,
            aca_nr_real_m2, aca_nr_beneficios_m2
        ) VALUES (
            %(codigo)s, %(numero_pa)s, %(tipo_processo)s, %(data_autuacao)s, %(status_pa)s,
            %(interessado)s, %(cnpj)s, %(cpf)s, %(tipo_interessado)s, %(endereco)s,
            %(setor_id)s::uuid, %(requerimento)s,
            %(contribuinte_sq)s, %(contribuinte_lote)s, %(area_terreno_m2)s,
            %(uso_aca)s, %(aca_r_m2)s, %(aca_nr_m2)s, %(aca_total_m2)s,
            %(tipo_contrapartida)s, %(valor_oodc_rs)s,
            %(cepac_aca)s, %(cepac_parametros)s, %(cepac_total)s,
            %(certidao)s, %(situacao_certidao)s, %(data_certidao)s,
            %(nuvem_r_m2)s, %(nuvem_nr_m2)s, %(nuvem_total_m2)s, %(nuvem_cepac)s,
            %(obs)s, %(resp_data)s, %(data_proposta)s,
            %(lei_vigente)s,
            %(aca_r_real_m2)s, %(aca_r_beneficios_m2)s,
            %(aca_nr_real_m2)s, %(aca_nr_beneficios_m2)s
        )
        ON CONFLICT (codigo) DO UPDATE SET
            numero_pa           = EXCLUDED.numero_pa,
            tipo_processo       = EXCLUDED.tipo_processo,
            data_autuacao       = EXCLUDED.data_autuacao,
            status_pa           = EXCLUDED.status_pa,
            interessado         = EXCLUDED.interessado,
            cnpj                = EXCLUDED.cnpj,
            cpf                 = EXCLUDED.cpf,
            tipo_interessado    = EXCLUDED.tipo_interessado,
            endereco            = EXCLUDED.endereco,
            setor_id            = EXCLUDED.setor_id,
            requerimento        = EXCLUDED.requerimento,
            contribuinte_sq     = EXCLUDED.contribuinte_sq,
            contribuinte_lote   = EXCLUDED.contribuinte_lote,
            area_terreno_m2     = EXCLUDED.area_terreno_m2,
            uso_aca             = EXCLUDED.uso_aca,
            aca_r_m2            = EXCLUDED.aca_r_m2,
            aca_nr_m2           = EXCLUDED.aca_nr_m2,
            aca_total_m2        = EXCLUDED.aca_total_m2,
            tipo_contrapartida  = EXCLUDED.tipo_contrapartida,
            valor_oodc_rs       = EXCLUDED.valor_oodc_rs,
            cepac_aca           = EXCLUDED.cepac_aca,
            cepac_parametros    = EXCLUDED.cepac_parametros,
            cepac_total         = EXCLUDED.cepac_total,
            certidao            = EXCLUDED.certidao,
            situacao_certidao   = EXCLUDED.situacao_certidao,
            data_certidao       = EXCLUDED.data_certidao,
            nuvem_r_m2          = EXCLUDED.nuvem_r_m2,
            nuvem_nr_m2         = EXCLUDED.nuvem_nr_m2,
            nuvem_total_m2      = EXCLUDED.nuvem_total_m2,
            nuvem_cepac         = EXCLUDED.nuvem_cepac,
            obs                 = EXCLUDED.obs,
            resp_data           = EXCLUDED.resp_data,
            data_proposta       = EXCLUDED.data_proposta,
            lei_vigente         = EXCLUDED.lei_vigente,
            aca_r_real_m2       = EXCLUDED.aca_r_real_m2,
            aca_r_beneficios_m2 = EXCLUDED.aca_r_beneficios_m2,
            aca_nr_real_m2      = EXCLUDED.aca_nr_real_m2,
            aca_nr_beneficios_m2 = EXCLUDED.aca_nr_beneficios_m2,
            updated_at          = now()
        RETURNING id, xmax
    """

    cnt_prop_insert = cnt_prop_update = cnt_prop_skip = 0
    erros_prop: list[str] = []
    proposta_id_map: dict[str, str] = {}  # codigo → UUID

    for codigo, linhas in sorted(grupos.items()):
        rep = selecionar_representativa(linhas)

        setor_nome_raw = str(rep.get("SETOR OUC") or "").strip()
        setor_id = setor_map.get(setor_nome_raw.upper())
        if not setor_id:
            msg = f"codigo={codigo!r}: setor não encontrado → {setor_nome_raw!r}"
            log.warning("SKIP proposta  %s", msg)
            erros_prop.append(msg)
            cnt_prop_skip += 1
            continue

        numero_pa, tipo_processo = detect_tipo_processo(
            rep.get("PA (16 digitos)"), rep.get("PA (12 digitos)")
        )
        cnpj = clean_semicolon(rep.get("CNPJ"))
        cpf = clean_semicolon(rep.get("CPF"))
        tipo_interessado = "PJ" if cnpj else ("PF" if cpf else None)

        req_raw = str(rep.get("REQUERIMENTO") or "").strip().upper()
        requerimento = REQUERIMENTO_MAP.get(req_raw)
        if not requerimento:
            msg = f"codigo={codigo!r}: requerimento inválido → {rep.get('REQUERIMENTO')!r}"
            log.warning("SKIP proposta  %s", msg)
            erros_prop.append(msg)
            cnt_prop_skip += 1
            continue

        status_raw = str(rep.get("STATUS PA") or "").strip().upper()
        status_pa = STATUS_PA_MAP.get(status_raw, "ANALISE")

        situacao_raw = str(rep.get("SITUACAO") or "").strip().upper()
        situacao = SITUACAO_MAP.get(situacao_raw)

        lei_raw = or_none(rep.get("LEI"))

        # ACA areas (net and real/benefícios)
        aca_r_real = parse_decimal(rep.get("ACA-R (REAL)"))
        aca_r_ben = parse_decimal(rep.get("ACA-R BENEFICIOS"))
        aca_nr_real = parse_decimal(rep.get("ACA-NR (REAL)"))
        aca_nr_ben = parse_decimal(rep.get("ACA-NR (BENEFÍCIOS)"))
        aca_r_net = compute_net_aca(rep.get("ACA-R"), rep.get("ACA-R (REAL)"), rep.get("ACA-R BENEFICIOS"))
        aca_nr_net = compute_net_aca(rep.get("ACA-NR"), rep.get("ACA-NR (REAL)"), rep.get("ACA-NR (BENEFÍCIOS)"))
        aca_total = parse_decimal(rep.get("ACA-TOTAL"))

        params = {
            "codigo": codigo,
            "numero_pa": numero_pa,
            "tipo_processo": tipo_processo,
            "data_autuacao": parse_date(rep.get("DATA AUTUACAO PA")),
            "status_pa": status_pa,
            "interessado": or_none(rep.get("INTERESSADO")),
            "cnpj": cnpj,
            "cpf": cpf,
            "tipo_interessado": tipo_interessado,
            "endereco": or_none(rep.get("ENDEREÇO")),
            "setor_id": setor_id,
            "requerimento": requerimento,
            "contribuinte_sq": clean_semicolon(rep.get("CONTRIBUINTE - SQ")),
            "contribuinte_lote": clean_semicolon(rep.get("CONTRIBUINTE - L")),
            "area_terreno_m2": parse_decimal(rep.get("ÁREA DO TERRENO")),
            "uso_aca": or_none(rep.get("USO DA ACA")),
            "aca_r_m2": aca_r_net,
            "aca_nr_m2": aca_nr_net,
            "aca_total_m2": aca_total,
            "tipo_contrapartida": or_none(rep.get("TIPO DE CONTRAPARTIDA")),
            "valor_oodc_rs": parse_decimal(rep.get("VALOR R$")),
            "cepac_aca": parse_integer(rep.get("CEPAC ACA")),
            "cepac_parametros": parse_integer(rep.get("CEPAC PARAMETROS")),
            "cepac_total": parse_integer(rep.get("TOTAL CEPAC")),
            "certidao": or_none(rep.get("CERTIDAO")),
            "situacao_certidao": situacao,
            "data_certidao": rep.get("_data_cert"),
            "nuvem_r_m2": parse_decimal(rep.get("NUVEM m² - R")),
            "nuvem_nr_m2": parse_decimal(rep.get("NUVEM m² -nR")),
            "nuvem_total_m2": parse_decimal(rep.get("NUVEM m²")),
            "nuvem_cepac": parse_integer(rep.get("NUVEM CEPAC")),
            "obs": or_none(rep.get("OBS")),
            "resp_data": or_none(rep.get("RESP/DATA")),
            "data_proposta": rep.get("_data_cert"),
            "lei_vigente": lei_raw,
            "aca_r_real_m2": aca_r_real,
            "aca_r_beneficios_m2": aca_r_ben,
            "aca_nr_real_m2": aca_nr_real,
            "aca_nr_beneficios_m2": aca_nr_ben,
        }

        if dry_run:
            proposta_id_map[codigo] = str(uuid4())
            cnt_prop_insert += 1
            continue

        try:
            cur.execute(UPSERT_PROPOSTA, params)
            result = cur.fetchone()
            proposta_id_map[codigo] = str(result["id"])
            if result["xmax"] == 0:
                cnt_prop_insert += 1
            else:
                cnt_prop_update += 1
        except Exception as exc:
            conn.rollback()
            msg = f"codigo={codigo!r}: erro SQL proposta → {exc}"
            log.error(msg)
            erros_prop.append(msg)
            cnt_prop_skip += 1

    if not dry_run:
        conn.commit()

    log.info(
        "Fase 1: %d inseridas, %d atualizadas, %d puladas",
        cnt_prop_insert, cnt_prop_update, cnt_prop_skip,
    )

    # -----------------------------------------------------------------------
    # FASE 2 — CERTIDÃO
    # -----------------------------------------------------------------------
    UPSERT_CERTIDAO = """
        INSERT INTO certidao (
            proposta_id, numero_certidao, tipo, data_emissao, situacao,
            numero_processo_sei, uso_aca,
            aca_r_m2, aca_nr_m2, aca_total_m2,
            tipo_contrapartida, valor_oodc_rs,
            cepac_aca, cepac_parametros, cepac_total,
            nuvem_r_m2, nuvem_nr_m2, nuvem_total_m2, nuvem_cepac,
            contribuinte_sq, contribuinte_lote, obs
        ) VALUES (
            %(proposta_id)s::uuid, %(numero_certidao)s, %(tipo)s::tipo_certidao_enum,
            %(data_emissao)s, %(situacao)s::situacao_certidao_enum,
            %(numero_processo_sei)s, %(uso_aca)s,
            %(aca_r_m2)s, %(aca_nr_m2)s, %(aca_total_m2)s,
            %(tipo_contrapartida)s, %(valor_oodc_rs)s,
            %(cepac_aca)s, %(cepac_parametros)s, %(cepac_total)s,
            %(nuvem_r_m2)s, %(nuvem_nr_m2)s, %(nuvem_total_m2)s, %(nuvem_cepac)s,
            %(contribuinte_sq)s, %(contribuinte_lote)s, %(obs)s
        )
        ON CONFLICT (numero_certidao) DO UPDATE SET
            proposta_id         = EXCLUDED.proposta_id,
            tipo                = EXCLUDED.tipo,
            data_emissao        = EXCLUDED.data_emissao,
            situacao            = EXCLUDED.situacao,
            numero_processo_sei = EXCLUDED.numero_processo_sei,
            uso_aca             = EXCLUDED.uso_aca,
            aca_r_m2            = EXCLUDED.aca_r_m2,
            aca_nr_m2           = EXCLUDED.aca_nr_m2,
            aca_total_m2        = EXCLUDED.aca_total_m2,
            tipo_contrapartida  = EXCLUDED.tipo_contrapartida,
            valor_oodc_rs       = EXCLUDED.valor_oodc_rs,
            cepac_aca           = EXCLUDED.cepac_aca,
            cepac_parametros    = EXCLUDED.cepac_parametros,
            cepac_total         = EXCLUDED.cepac_total,
            nuvem_r_m2          = EXCLUDED.nuvem_r_m2,
            nuvem_nr_m2         = EXCLUDED.nuvem_nr_m2,
            nuvem_total_m2      = EXCLUDED.nuvem_total_m2,
            nuvem_cepac         = EXCLUDED.nuvem_cepac,
            contribuinte_sq     = EXCLUDED.contribuinte_sq,
            contribuinte_lote   = EXCLUDED.contribuinte_lote,
            obs                 = EXCLUDED.obs
        WHERE certidao.situacao != 'VALIDA'::situacao_certidao_enum
           OR EXCLUDED.situacao = 'VALIDA'::situacao_certidao_enum
        RETURNING xmax
    """

    cnt_cert_insert = cnt_cert_update = cnt_cert_skip = 0
    erros_cert: list[str] = []

    # Para validação: totais por lei
    totais: dict[str, dict[str, Decimal]] = defaultdict(lambda: {
        "r_consumido": Decimal("0"), "nr_consumido": Decimal("0"),
        "r_analise": Decimal("0"), "nr_analise": Decimal("0"),
    })

    # Ordena linhas: CANCELADA primeiro, ANALISE depois, VALIDA por último (VALIDA vence conflito)
    rows_sorted = sorted(
        rows_enriched,
        key=lambda r: SITUACAO_PRIORIDADE.get(
            str(r.get("SITUACAO") or "").strip().upper(), 99
        ),
        reverse=True,  # CANCELADA(2) → ANALISE(1) → VALIDA(0) processada por último
    )

    for linha_idx, row in enumerate(rows_sorted):
        codigo = str(row.get("CODIGO") or "").strip()
        proposta_id = proposta_id_map.get(codigo)
        if not proposta_id:
            cnt_cert_skip += 1
            continue

        req_raw = str(row.get("REQUERIMENTO") or "").strip().upper()
        tipo_cert = TIPO_CERTIDAO_MAP.get(req_raw, "VINCULAÇÃO")

        situacao_raw = str(row.get("SITUACAO") or "").strip().upper()
        situacao = SITUACAO_MAP.get(situacao_raw, "ANALISE")

        numero_certidao_raw = or_none(row.get("CERTIDAO"))
        if numero_certidao_raw:
            numero_certidao = numero_certidao_raw[:20]
        else:
            numero_certidao = f"FL-{codigo}-{linha_idx}"[:20]

        numero_pa, _ = detect_tipo_processo(
            row.get("PA (16 digitos)"), row.get("PA (12 digitos)")
        )

        aca_r_net = compute_net_aca(row.get("ACA-R"), row.get("ACA-R (REAL)"), row.get("ACA-R BENEFICIOS"))
        aca_nr_net = compute_net_aca(row.get("ACA-NR"), row.get("ACA-NR (REAL)"), row.get("ACA-NR (BENEFÍCIOS)"))

        # Acumula totais para relatório de validação
        lei = or_none(row.get("LEI")) or "NULL"
        status_raw = str(row.get("STATUS PA") or "").strip().upper()
        if status_raw == "DEFERIDO" and situacao_raw == "VALIDA":
            totais[lei]["r_consumido"] += aca_r_net or Decimal("0")
            totais[lei]["nr_consumido"] += aca_nr_net or Decimal("0")
        elif status_raw == "ANALISE":
            totais[lei]["r_analise"] += aca_r_net or Decimal("0")
            totais[lei]["nr_analise"] += aca_nr_net or Decimal("0")

        params = {
            "proposta_id": proposta_id,
            "numero_certidao": numero_certidao,
            "tipo": tipo_cert,
            "data_emissao": row.get("_data_cert"),
            "situacao": situacao,
            "numero_processo_sei": numero_pa,
            "uso_aca": or_none(row.get("USO DA ACA")),
            "aca_r_m2": aca_r_net,
            "aca_nr_m2": aca_nr_net,
            "aca_total_m2": parse_decimal(row.get("ACA-TOTAL")),
            "tipo_contrapartida": or_none(row.get("TIPO DE CONTRAPARTIDA")),
            "valor_oodc_rs": parse_decimal(row.get("VALOR R$")),
            "cepac_aca": parse_integer(row.get("CEPAC ACA")),
            "cepac_parametros": parse_integer(row.get("CEPAC PARAMETROS")),
            "cepac_total": parse_integer(row.get("TOTAL CEPAC")),
            "nuvem_r_m2": parse_decimal(row.get("NUVEM m² - R")),
            "nuvem_nr_m2": parse_decimal(row.get("NUVEM m² -nR")),
            "nuvem_total_m2": parse_decimal(row.get("NUVEM m²")),
            "nuvem_cepac": parse_integer(row.get("NUVEM CEPAC")),
            "contribuinte_sq": clean_semicolon(row.get("CONTRIBUINTE - SQ")),
            "contribuinte_lote": clean_semicolon(row.get("CONTRIBUINTE - L")),
            "obs": or_none(row.get("OBS")),
        }

        if dry_run:
            cnt_cert_insert += 1
            continue

        try:
            cur.execute(UPSERT_CERTIDAO, params)
            result = cur.fetchone()
            if result is None:
                # DO UPDATE WHERE condition was not met — skip
                cnt_cert_skip += 1
            elif result["xmax"] == 0:
                cnt_cert_insert += 1
            else:
                cnt_cert_update += 1
        except Exception as exc:
            conn.rollback()
            msg = f"codigo={codigo!r} certidao={numero_certidao!r}: erro SQL → {exc}"
            log.error(msg)
            erros_cert.append(msg)
            cnt_cert_skip += 1

    if not dry_run:
        conn.commit()

    log.info(
        "Fase 2: %d inseridas, %d atualizadas, %d puladas",
        cnt_cert_insert, cnt_cert_update, cnt_cert_skip,
    )

    # -----------------------------------------------------------------------
    # FASE 3 — TituloCepac + Movimentacao
    # -----------------------------------------------------------------------
    UPSERT_TITULO = """
        INSERT INTO titulo_cepac (id, codigo, setor_id, valor_m2, uso, origem, estado)
        VALUES (%(id)s::uuid, %(codigo)s, %(setor_id)s::uuid, %(valor_m2)s,
                %(uso)s::uso_enum, %(origem)s::origem_enum, %(estado)s::estado_titulo_enum)
        ON CONFLICT (codigo) DO NOTHING
        RETURNING id
    """
    INSERT_MOVIMENTACAO = """
        INSERT INTO movimentacao (
            titulo_id, setor_id, uso, origem,
            estado_anterior, estado_novo,
            numero_processo_sei, motivo, incentivado, operador
        ) VALUES (
            %(titulo_id)s::uuid, %(setor_id)s::uuid,
            %(uso)s::uso_enum, %(origem)s::origem_enum,
            NULL, %(estado_novo)s::estado_titulo_enum,
            %(numero_processo_sei)s, %(motivo)s, %(incentivado)s, %(operador)s
        )
    """

    cnt_titulo_13769 = cnt_titulo_13769_skip = 0
    cnt_titulo_18175 = cnt_titulo_18175_skip = 0
    erros_titulo: list[str] = []

    def _inserir_titulo(setor_nome_upper: str, uso: str, valor_m2: Decimal,
                        estado: str, numero_processo_sei: str, motivo: str,
                        titulo_codigo: str) -> bool:
        """Retorna True se inserido, False se já existe (skip)."""
        setor_id = setor_map.get(setor_nome_upper)
        if not setor_id:
            return False
        titulo_id = str(uuid4())
        cur.execute(UPSERT_TITULO, {
            "id": titulo_id,
            "codigo": titulo_codigo[:50],
            "setor_id": setor_id,
            "valor_m2": str(valor_m2),
            "uso": uso,
            "origem": "ACA",
            "estado": estado,
        })
        result = cur.fetchone()
        if result is None:
            return False  # ON CONFLICT DO NOTHING
        cur.execute(INSERT_MOVIMENTACAO, {
            "titulo_id": result["id"],
            "setor_id": setor_id,
            "uso": uso,
            "origem": "ACA",
            "estado_novo": estado,
            "numero_processo_sei": numero_processo_sei[:50],
            "motivo": motivo,
            "incentivado": False,
            "operador": OPERADOR_SCRIPT,
        })
        return True

    # --- FASE 3A: consumo histórico consolidado Lei 13.769/2004 ---
    for setor_upper, dims in HISTORICO_13769.items():
        for uso, valor_m2 in dims.items():
            if valor_m2 <= 0:
                continue
            titulo_codigo = f"OUCFL-13769-{setor_upper[:12].replace(' ', '')}-{uso}"

            if dry_run:
                cnt_titulo_13769 += 1
                continue
            try:
                inserido = _inserir_titulo(
                    setor_nome_upper=setor_upper,
                    uso=uso,
                    valor_m2=valor_m2,
                    estado="CONSUMIDO",
                    numero_processo_sei="CARGA_HISTORICA_13769",
                    motivo=MOTIVO_CARGA_13769,
                    titulo_codigo=titulo_codigo,
                )
                if inserido:
                    cnt_titulo_13769 += 1
                else:
                    cnt_titulo_13769_skip += 1
            except Exception as exc:
                conn.rollback()
                erros_titulo.append(f"13769 {setor_upper} {uso}: {exc}")
                cnt_titulo_13769_skip += 1

    if not dry_run:
        conn.commit()

    log.info(
        "Fase 3A (TituloCepac 13.769 histórico): %d inseridos, %d já existiam",
        cnt_titulo_13769, cnt_titulo_13769_skip,
    )

    # --- FASE 3B: certidões individuais Lei 18.175/2024 ---
    for row in rows_enriched:
        lei = or_none(row.get("LEI")) or ""
        if LEI_VIGENTE not in lei:
            continue

        req_raw = str(row.get("REQUERIMENTO") or "").strip().upper()
        if req_raw not in ("VINCULACAO", "ALTERACAO"):
            continue

        status_raw = str(row.get("STATUS PA") or "").strip().upper()
        situacao_raw = str(row.get("SITUACAO") or "").strip().upper()

        if status_raw == "DEFERIDO" and situacao_raw == "VALIDA":
            estado = "CONSUMIDO"
        elif situacao_raw == "ANALISE":
            estado = "EM_ANALISE"
        else:
            continue

        codigo = str(row.get("CODIGO") or "").strip()
        setor_nome = str(row.get("SETOR OUC") or "").strip()
        if not setor_map.get(setor_nome.upper()):
            continue

        certidao_num = or_none(row.get("CERTIDAO")) or codigo

        numero_pa, _ = detect_tipo_processo(
            row.get("PA (16 digitos)"), row.get("PA (12 digitos)")
        )
        numero_processo_sei = numero_pa or codigo[:50]

        aca_r_net = compute_net_aca(row.get("ACA-R"), row.get("ACA-R (REAL)"), row.get("ACA-R BENEFICIOS"))
        aca_nr_net = compute_net_aca(row.get("ACA-NR"), row.get("ACA-NR (REAL)"), row.get("ACA-NR (BENEFÍCIOS)"))

        dimensoes = []
        if aca_r_net and aca_r_net > 0:
            dimensoes.append(("R", aca_r_net))
        if aca_nr_net and aca_nr_net > 0:
            dimensoes.append(("NR", aca_nr_net))

        for uso, valor_m2 in dimensoes:
            titulo_codigo = f"OUCFL-{certidao_num}-{uso}"

            if dry_run:
                cnt_titulo_18175 += 1
                continue

            try:
                inserido = _inserir_titulo(
                    setor_nome_upper=setor_nome.upper(),
                    uso=uso,
                    valor_m2=valor_m2,
                    estado=estado,
                    numero_processo_sei=numero_processo_sei,
                    motivo=MOTIVO_CARGA_18175,
                    titulo_codigo=titulo_codigo,
                )
                if inserido:
                    cnt_titulo_18175 += 1
                else:
                    cnt_titulo_18175_skip += 1
            except Exception as exc:
                conn.rollback()
                msg = f"titulo={titulo_codigo!r}: {exc}"
                log.error(msg)
                erros_titulo.append(msg)
                cnt_titulo_18175_skip += 1

    if not dry_run:
        conn.commit()

    log.info(
        "Fase 3B (TituloCepac 18.175 individual): %d inseridos, %d já existiam",
        cnt_titulo_18175, cnt_titulo_18175_skip,
    )

    conn.close()

    # -----------------------------------------------------------------------
    # Relatório final
    # -----------------------------------------------------------------------
    print()
    print("=" * 65)
    print("  RESULTADO DA CARGA OUCFL" + ("  [DRY-RUN — nenhum dado gravado]" if dry_run else ""))
    print("=" * 65)
    print()
    print("  PROPOSTAS")
    print(f"    Inseridas                  : {cnt_prop_insert:>6}")
    print(f"    Atualizadas                : {cnt_prop_update:>6}")
    print(f"    Puladas/erro               : {cnt_prop_skip:>6}")
    print()
    print("  CERTIDÕES")
    print(f"    Inseridas                  : {cnt_cert_insert:>6}")
    print(f"    Atualizadas                : {cnt_cert_update:>6}")
    print(f"    Puladas                    : {cnt_cert_skip:>6}")
    print()
    print("  TITULO_CEPAC + MOVIMENTACAO (Lei 13.769/2004 — histórico batch)")
    print(f"    Inseridos                  : {cnt_titulo_13769:>6}")
    print(f"    Já existiam (skip)         : {cnt_titulo_13769_skip:>6}")
    print()
    print("  TITULO_CEPAC + MOVIMENTACAO (Lei 18.175/2024 — individual)")
    print(f"    Inseridos                  : {cnt_titulo_18175:>6}")
    print(f"    Já existiam (skip)         : {cnt_titulo_18175_skip:>6}")
    print()
    print("  TOTAIS POR LEI (para validação)")
    print("  {:<20} {:>12} {:>12} {:>12} {:>12}".format(
        "LEI", "R Consumido", "NR Consumido", "R Análise", "NR Análise"
    ))
    print("  " + "-" * 62)
    for lei, t in sorted(totais.items()):
        print("  {:<20} {:>12.2f} {:>12.2f} {:>12.2f} {:>12.2f}".format(
            lei, t["r_consumido"], t["nr_consumido"], t["r_analise"], t["nr_analise"]
        ))
    print()
    print("  ESPERADO (planilha Consolidado_OUC-FL):")
    print("  Lei 13.769/2004: R=646099.80  NR=426934.24  (congelado em lei_ouc.consumo_historico)")
    print("  Lei 18.175/2024: R=70352.02   NR=51046.38   (após benefícios — benefícios")
    print("                   incompletos na aba 2_CONTROLE_ESTOQUE)")
    print("  Em análise 18.175: R=4854.71  NR=175.83")
    print()
    if erros_prop or erros_cert or erros_titulo:
        print("  ERROS DETALHADOS:")
        for e in erros_prop + erros_cert + erros_titulo:
            print(f"    - {e}")
        print()
    print("=" * 65)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Carga OUCFL no banco CEPAC")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Calcula e exibe os totais sem gravar no banco",
    )
    args = parser.parse_args()

    load_env(ENV_PATH)

    if not XLSX_PATH.exists():
        log.error("Planilha não encontrada: %s", XLSX_PATH)
        sys.exit(1)

    try:
        connect_kwargs = get_connect_kwargs()
    except RuntimeError as e:
        log.error(str(e))
        sys.exit(1)

    log.info("Planilha: %s", XLSX_PATH)
    log.info("Banco: %s@%s/%s", connect_kwargs.get("user"), connect_kwargs.get("host"), connect_kwargs.get("dbname"))

    carregar(connect_kwargs, XLSX_PATH, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
