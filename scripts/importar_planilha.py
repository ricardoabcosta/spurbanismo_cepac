#!/usr/bin/env python3
"""
scripts/importar_planilha.py
SP Urbanismo / Prodam — CEPAC T11

Lê docs/OUCAE_ESTOQUE_abr_rv01.xlsx e gera dois arquivos SQL:
  - migrations/004a_revogar_seed_sintetico.sql (remove SEED-* da migration 002)
  - migrations/004_carga_inicial_real.sql       (insert dos dados reais)

Uso:
    python scripts/importar_planilha.py [--xlsx <path>]

Saída:
  - migrations/004a_revogar_seed_sintetico.sql
  - migrations/004_carga_inicial_real.sql
  - imprime resumo de validação para conferência com Consolidado_OUC-AE

Regras de mapeamento (2_CONTROLE_ESTOQUE):
  - DEFERIDO + VALIDA + VINCULACAO/ALTERACAO  → estado CONSUMIDO
  - DEFERIDO + VALIDA + DESVINCULACAO          → DISPONIVEL (≥ 180 dias) ou QUARENTENA (< 180 dias)
  - ANALISE  + *      + *                      → EM_ANALISE
  - INDEFERIDO                                 → ignorado
  - DEFERIDO + CANCELADA                       → ignorado

Cada linha que passa nos filtros gera:
  1. INSERT proposta    (ON CONFLICT DO NOTHING — um por código único)
  2. INSERT certidao    (ON CONFLICT DO NOTHING — uma por numero_certidao)
  3. INSERT titulo_cepac (uma por combinação uso×origem com área > 0)
  4. INSERT movimentacao (uma por titulo_cepac)

Certidões da ABA 3 (3_controle_certidoes) são acrescentadas ao final,
vinculadas às propostas já inseridas.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl não instalado. Execute: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

TODAY = date(2026, 4, 16)
QUARENTENA_DIAS = 180

SETOR_MAP = {
    "BROOKLIN": "Brooklin",
    "BERRINI": "Berrini",
    "MARGINAL PINHEIROS": "Marginal Pinheiros",
    "CHUCRI ZAIDAN": "Chucri Zaidan",
    "JABAQUARA": "Jabaquara",
}

TIPO_CERTIDAO_MAP = {
    "VINCULAÇÃO": "VINCULAÇÃO",
    "DESVINCULAÇÃO": "DESVINCULAÇÃO",
    "ALTERAÇÃO": "ALTERAÇÃO",
    # Fallback sem acento (defensivo)
    "VINCULACAO": "VINCULAÇÃO",
    "DESVINCULACAO": "DESVINCULAÇÃO",
    "ALTERACAO": "ALTERAÇÃO",
}

TIPO_CERTIDAO_FROM_REQ = {
    "VINCULACAO": "VINCULAÇÃO",
    "DESVINCULACAO": "DESVINCULAÇÃO",
    "ALTERACAO": "ALTERAÇÃO",
}

# ---------------------------------------------------------------------------
# Utilitários SQL
# ---------------------------------------------------------------------------

def sql_str(v: Any) -> str:
    """Escapa string para SQL (comillas simples)."""
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

def sql_date(v: Any) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        v = v.date()
    return f"'{v.isoformat()}'"

def sql_num(v: Any, decimals: int = 2) -> str:
    if v is None or v == 0:
        return "NULL"
    return str(Decimal(str(v)).quantize(Decimal(10) ** -decimals, rounding=ROUND_HALF_UP))


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------

def load_workbook(xlsx_path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(str(xlsx_path), data_only=True)


def parse_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def map_estado(status_pa: str, situacao: str | None, req: str, data_cert: date | None) -> str:
    """
    Determina o estado do título baseado no registro da planilha.

    ANALISE + qualquer situação → EM_ANALISE
    DEFERIDO + VALIDA + DESVINCULACAO → DISPONIVEL ou QUARENTENA por data
    DEFERIDO + VALIDA + outros → CONSUMIDO
    """
    if status_pa == "ANALISE":
        return "EM_ANALISE"
    # DEFERIDO + VALIDA
    if req == "DESVINCULACAO":
        if data_cert is not None and (TODAY - data_cert).days >= QUARENTENA_DIAS:
            return "DISPONIVEL"
        return "QUARENTENA"
    return "CONSUMIDO"


def should_import(status_pa: str | None, situacao: str | None, req: str | None, setor: str | None) -> bool:
    if not status_pa or not req or not setor:
        return False
    if setor not in SETOR_MAP:
        return False
    if status_pa == "INDEFERIDO":
        return False
    if status_pa == "DEFERIDO" and situacao == "CANCELADA":
        return False
    return True


def pa_info(pa16: Any, pa12: Any) -> tuple[str | None, str | None]:
    """Retorna (numero_pa, tipo_processo)."""
    if pa16:
        return str(pa16), "SEI"
    if pa12:
        return str(pa12), "SIMPROC"
    return None, None


# ---------------------------------------------------------------------------
# Geração dos INSERTs
# ---------------------------------------------------------------------------

def gerar_insert_proposta(
    codigo: str,
    numero_pa: str | None,
    tipo_processo: str | None,
    data_autuacao: date | None,
    status_pa: str,
    interessado: str | None,
    cnpj_cpf: str | None,
    endereco: str | None,
    setor_nome: str,
    requerimento: str,
    area_terreno: Any,
) -> str:
    return (
        f"INSERT INTO proposta (codigo, numero_pa, tipo_processo, data_autuacao, "
        f"status_pa, interessado, cnpj_cpf, endereco, setor_id, requerimento, area_terreno_m2)\n"
        f"SELECT\n"
        f"    {sql_str(codigo)},\n"
        f"    {sql_str(numero_pa)},\n"
        f"    {sql_str(tipo_processo)}::tipo_processo_enum,\n"
        f"    {sql_date(data_autuacao)},\n"
        f"    {sql_str(status_pa)}::status_pa_enum,\n"
        f"    {sql_str(interessado)},\n"
        f"    {sql_str(cnpj_cpf)},\n"
        f"    {sql_str(endereco)},\n"
        f"    (SELECT id FROM setor WHERE nome = {sql_str(setor_nome)}),\n"
        f"    {sql_str(requerimento)}::requerimento_enum,\n"
        f"    {sql_num(area_terreno)}\n"
        f"ON CONFLICT (codigo) DO NOTHING;"
    )


def gerar_insert_certidao(
    proposta_codigo: str,
    numero_certidao: str,
    tipo: str,
    data_emissao: date | None,
    numero_processo_sei: str | None,
    situacao: str,
) -> str:
    return (
        f"INSERT INTO certidao (proposta_id, numero_certidao, tipo, data_emissao, numero_processo_sei, situacao)\n"
        f"SELECT\n"
        f"    (SELECT id FROM proposta WHERE codigo = {sql_str(proposta_codigo)}),\n"
        f"    {sql_str(numero_certidao)},\n"
        f"    {sql_str(tipo)}::tipo_certidao_enum,\n"
        f"    {sql_date(data_emissao)},\n"
        f"    {sql_str(numero_processo_sei)},\n"
        f"    {sql_str(situacao)}::situacao_certidao_enum\n"
        f"ON CONFLICT (numero_certidao) DO NOTHING;"
    )


def gerar_insert_titulo_e_mov(
    codigo_titulo: str,
    setor_nome: str,
    uso: str,
    origem: str,
    valor_m2: float,
    estado: str,
    data_desvinculacao: date | None,
    numero_processo_sei: str,
) -> list[str]:
    """Gera INSERT de titulo_cepac e INSERT de movimentacao para um título."""
    data_desv_sql = sql_date(data_desvinculacao) if estado in ("QUARENTENA", "DISPONIVEL") else "NULL"

    titulo_sql = (
        f"INSERT INTO titulo_cepac (codigo, setor_id, valor_m2, uso, origem, estado, data_desvinculacao)\n"
        f"SELECT\n"
        f"    {sql_str(codigo_titulo)},\n"
        f"    (SELECT id FROM setor WHERE nome = {sql_str(setor_nome)}),\n"
        f"    {sql_num(valor_m2)},\n"
        f"    '{uso}'::uso_enum,\n"
        f"    '{origem}'::origem_enum,\n"
        f"    '{estado}'::estado_titulo_enum,\n"
        f"    {data_desv_sql}\n"
        f"WHERE NOT EXISTS (SELECT 1 FROM titulo_cepac WHERE codigo = {sql_str(codigo_titulo)});"
    )

    mov_sql = (
        f"INSERT INTO movimentacao (titulo_id, setor_id, uso, origem, estado_anterior, estado_novo, "
        f"numero_processo_sei, motivo, operador)\n"
        f"SELECT\n"
        f"    t.id, t.setor_id, t.uso, t.origem,\n"
        f"    NULL,\n"
        f"    '{estado}'::estado_titulo_enum,\n"
        f"    {sql_str(numero_processo_sei)},\n"
        f"    'IMPORTACAO_XLSX',\n"
        f"    'SISTEMA'\n"
        f"FROM titulo_cepac t\n"
        f"WHERE t.codigo = {sql_str(codigo_titulo)}\n"
        f"  AND NOT EXISTS (\n"
        f"      SELECT 1 FROM movimentacao m\n"
        f"      WHERE m.titulo_id = t.id AND m.motivo = 'IMPORTACAO_XLSX'\n"
        f"  );"
    )

    return [titulo_sql, mov_sql]


def extrair_areas(row: tuple) -> list[tuple[str, str, float]]:
    """
    Retorna lista de (uso, origem, valor_m2) para cada área não-nula do registro.

    ACA-R  → (R, ACA, valor)
    ACA-NR → (NR, ACA, valor)
    NUVEM-R  → (R, NUVEM, valor)
    NUVEM-NR → (NR, NUVEM, valor)
    """
    areas = []
    aca_r = float(row[15] or 0)
    aca_nr = float(row[16] or 0)
    nuvem_r = float(row[26] or 0)
    nuvem_nr = float(row[27] or 0)

    if aca_r > 0:
        areas.append(("R", "ACA", aca_r))
    if aca_nr > 0:
        areas.append(("NR", "ACA", aca_nr))
    if nuvem_r > 0:
        areas.append(("R", "NUVEM", nuvem_r))
    if nuvem_nr > 0:
        areas.append(("NR", "NUVEM", nuvem_nr))

    return areas


# ---------------------------------------------------------------------------
# Processamento principal
# ---------------------------------------------------------------------------

def processar(xlsx_path: Path) -> tuple[list[str], dict]:
    """
    Lê a planilha e retorna (lista_de_sqls, resumo_validacao).
    """
    wb = load_workbook(xlsx_path)
    ws2 = wb["2_CONTROLE_ESTOQUE"]
    ws3 = wb["3_controle_certidoes"]

    sqls: list[str] = []

    # Controle de idempotência local
    codigos_proposta_vistos: set[str] = set()
    numeros_certidao_vistos: set[str] = set()
    codigos_titulo_vistos: set[str] = set()

    # Contador de títulos por código de proposta (para desambiguação)
    contador_titulos: dict[str, int] = defaultdict(int)

    # Validação: somas por (setor, uso, origem, estado)
    totais: dict = defaultdict(float)

    # -----------------------------------------------------------------------
    # Etapa 1: processar ABA 2_CONTROLE_ESTOQUE
    # -----------------------------------------------------------------------
    for row in ws2.iter_rows(min_row=2, values_only=True):
        codigo = row[0]
        if codigo is None:
            continue

        status_pa = row[4]
        situacao = row[24]
        req = row[10]
        setor_upper = row[9]

        if not should_import(status_pa, situacao, req, setor_upper):
            continue

        setor_nome = SETOR_MAP[setor_upper]
        pa16, pa12 = row[1], row[2]
        numero_pa, tipo_processo = pa_info(pa16, pa12)
        data_autuacao = parse_date(row[3])
        interessado = row[5]
        cnpj_raw = row[6]
        cpf_raw = row[7]
        cnpj_cpf = str(cnpj_raw or cpf_raw or "").strip() or None
        endereco = row[8]
        area_terreno = row[13]
        numero_certidao_raw = row[23]
        data_cert = parse_date(row[25])

        estado = map_estado(status_pa, situacao, req, data_cert)

        # 1a. Proposta (uma por código, ON CONFLICT DO NOTHING)
        if codigo not in codigos_proposta_vistos:
            codigos_proposta_vistos.add(codigo)
            sqls.append(
                gerar_insert_proposta(
                    codigo=codigo,
                    numero_pa=numero_pa,
                    tipo_processo=tipo_processo,
                    data_autuacao=data_autuacao,
                    status_pa=status_pa,
                    interessado=interessado if interessado else None,
                    cnpj_cpf=cnpj_cpf,
                    endereco=endereco if endereco else None,
                    setor_nome=setor_nome,
                    requerimento=req,
                    area_terreno=area_terreno,
                )
            )

        # 1b. Certidão (se coluna CERTIDAO preenchida)
        numero_certidao = str(numero_certidao_raw).strip() if numero_certidao_raw else None
        if numero_certidao and numero_certidao not in numeros_certidao_vistos:
            numeros_certidao_vistos.add(numero_certidao)
            tipo_cert = TIPO_CERTIDAO_FROM_REQ.get(req, "VINCULAÇÃO")
            # Situação da certidão na ABA 2: VALIDA (padrão) ou CANCELADA se situacao=CANCELADA
            # Neste ponto, apenas linhas com situacao != CANCELADA chegam aqui
            situacao_cert = "VALIDA"
            sqls.append(
                gerar_insert_certidao(
                    proposta_codigo=codigo,
                    numero_certidao=numero_certidao,
                    tipo=tipo_cert,
                    data_emissao=data_cert,
                    numero_processo_sei=numero_pa,
                    situacao=situacao_cert,
                )
            )

        # 1c. Títulos (um por área > 0)
        areas = extrair_areas(row)
        numero_sei_titulo = numero_pa or "IMPORTACAO-XLSX"

        for uso, origem, valor_m2 in areas:
            contador_titulos[codigo] += 1
            seq = contador_titulos[codigo]
            # Código do título: usa certidão se disponível, senão código da proposta + seq
            if numero_certidao:
                base = numero_certidao.replace("/", "-").replace(" ", "")
            else:
                base = f"{codigo}-S{seq:03d}"
            codigo_titulo = f"{base}-{uso}-{origem}"

            # Garantir unicidade (podem ocorrer colisões em edge cases)
            original_codigo = codigo_titulo
            dedup = 2
            while codigo_titulo in codigos_titulo_vistos:
                codigo_titulo = f"{original_codigo}-{dedup}"
                dedup += 1
            codigos_titulo_vistos.add(codigo_titulo)

            # data_desvinculacao para QUARENTENA e DISPONIVEL (vem de data certidão)
            data_desv: date | None = None
            if estado in ("QUARENTENA", "DISPONIVEL"):
                data_desv = data_cert

            stmts = gerar_insert_titulo_e_mov(
                codigo_titulo=codigo_titulo,
                setor_nome=setor_nome,
                uso=uso,
                origem=origem,
                valor_m2=valor_m2,
                estado=estado,
                data_desvinculacao=data_desv,
                numero_processo_sei=numero_sei_titulo,
            )
            sqls.extend(stmts)

            # Acumular totais para validação
            totais[(setor_nome, uso, origem, estado)] += valor_m2

    # -----------------------------------------------------------------------
    # Etapa 2: certidões da ABA 3_controle_certidoes (módulo consulta pública)
    # -----------------------------------------------------------------------
    for row in ws3.iter_rows(min_row=3, values_only=True):  # linha 1 = modelo/header
        if not any(v for v in row):
            continue
        tipo_raw = row[0]
        processo = row[1]
        data_emissao_raw = row[2]
        proposta_codigo = row[3]
        numero_certidao_aba3 = row[7]
        situacao_raw = row[8]

        if not numero_certidao_aba3 or not proposta_codigo:
            continue

        numero_certidao_aba3 = str(numero_certidao_aba3).strip()
        proposta_codigo = str(proposta_codigo).strip()

        # Ignorar linhas de modelo
        if proposta_codigo == "XX-0000":
            continue

        if numero_certidao_aba3 in numeros_certidao_vistos:
            continue  # já inserida pela ABA 2
        numeros_certidao_vistos.add(numero_certidao_aba3)

        tipo_cert = TIPO_CERTIDAO_MAP.get(str(tipo_raw or "").strip(), "VINCULAÇÃO")
        data_emissao = parse_date(data_emissao_raw)
        situacao = "CANCELADA" if str(situacao_raw or "").strip() == "CANCELADA" else "VALIDA"
        numero_sei = str(processo).strip() if processo else None

        sqls.append(
            gerar_insert_certidao(
                proposta_codigo=proposta_codigo,
                numero_certidao=numero_certidao_aba3,
                tipo=tipo_cert,
                data_emissao=data_emissao,
                numero_processo_sei=numero_sei,
                situacao=situacao,
            )
        )

    return sqls, totais


# ---------------------------------------------------------------------------
# Validação e impressão de resumo
# ---------------------------------------------------------------------------

EXPECTED_CONSUMIDO = {
    ("Brooklin", "R", "ACA"): 716469.73,
    ("Brooklin", "NR", "ACA"): 113342.09,
    ("Berrini", "R", "ACA"): 134590.11,
    ("Berrini", "NR", "ACA"): 202607.23,
    ("Berrini", "R", "NUVEM"): 100.62,
    ("Berrini", "NR", "NUVEM"): 595.00,
    ("Marginal Pinheiros", "R", "ACA"): 198945.31,
    ("Marginal Pinheiros", "NR", "ACA"): 258908.19,
    ("Marginal Pinheiros", "R", "NUVEM"): 1301.13,
    ("Chucri Zaidan", "R", "ACA"): 737902.74,
    ("Chucri Zaidan", "NR", "ACA"): 1050881.42,
    ("Chucri Zaidan", "R", "NUVEM"): 204.70,
    ("Chucri Zaidan", "NR", "NUVEM"): 434.79,
    ("Jabaquara", "R", "ACA"): 7709.85,
}

EXPECTED_EM_ANALISE = {
    ("Brooklin", "R", "ACA"): 8795.18,
    ("Chucri Zaidan", "R", "ACA"): 14006.35,
    ("Marginal Pinheiros", "R", "ACA"): 59398.03,
    ("Marginal Pinheiros", "NR", "ACA"): 11173.06,
}


def imprimir_validacao(totais: dict) -> bool:
    """
    Imprime resumo de validação. Retorna True se todos os valores batem.
    """
    ok = True
    print("\n" + "=" * 70)
    print("VALIDAÇÃO — CONSUMIDO (DEFERIDO+VALIDA)")
    print("=" * 70)
    for (setor, uso, origem), esperado in sorted(EXPECTED_CONSUMIDO.items()):
        obtido = totais.get((setor, uso, origem, "CONSUMIDO"), 0.0)
        diff = abs(obtido - esperado)
        status = "✓" if diff < 1.0 else "✗"
        if diff >= 1.0:
            ok = False
        print(f"  {status} {setor:22s} {uso:3s} {origem:6s} got={obtido:14.2f} expected={esperado:14.2f} diff={diff:.4f}")

    print("\n" + "=" * 70)
    print("VALIDAÇÃO — EM_ANALISE (ANALISE+*)")
    print("=" * 70)
    for (setor, uso, origem), esperado in sorted(EXPECTED_EM_ANALISE.items()):
        obtido = totais.get((setor, uso, origem, "EM_ANALISE"), 0.0)
        diff = abs(obtido - esperado)
        status = "✓" if diff < 1.0 else "✗"
        if diff >= 1.0:
            ok = False
        print(f"  {status} {setor:22s} {uso:3s} {origem:6s} got={obtido:14.2f} expected={esperado:14.2f} diff={diff:.4f}")

    # Totais de desvinculações
    print("\n" + "=" * 70)
    print("TOTAIS — QUARENTENA e DISPONIVEL (DESVINCULACAO)")
    print("=" * 70)
    for (setor, uso, origem, estado), total in sorted(totais.items()):
        if estado in ("QUARENTENA", "DISPONIVEL"):
            print(f"  {setor:22s} {uso:3s} {origem:6s} {estado:12s} {total:14.2f}")

    if ok:
        print("\n✓ TODOS OS TOTAIS CONFERIDOS. Migration pronta para revisão.")
    else:
        print("\n✗ DIVERGÊNCIAS ENCONTRADAS. Revisar antes de executar em produção.")

    return ok


# ---------------------------------------------------------------------------
# Escrita dos arquivos de saída
# ---------------------------------------------------------------------------

HEADER_004A = """\
-- =============================================================================
-- CEPAC — SP Urbanismo / Prodam
-- Migration 004a: Remoção do Seed Sintético (Migration 002)
-- PostgreSQL 15
--
-- Remove os 13 títulos SEED-* inseridos pela migration 002_seed_abril_2026.sql.
-- Deve ser executada ANTES de 004_carga_inicial_real.sql.
-- Idempotente: DELETE ... WHERE codigo LIKE 'SEED-%' não falha se já removido.
-- =============================================================================

BEGIN;

-- Remove movimentações dos títulos sintéticos
DELETE FROM movimentacao
WHERE titulo_id IN (
    SELECT id FROM titulo_cepac WHERE codigo LIKE 'SEED-%'
);

-- Remove os títulos sintéticos
DELETE FROM titulo_cepac WHERE codigo LIKE 'SEED-%';

COMMIT;
"""

HEADER_004 = """\
-- =============================================================================
-- CEPAC — SP Urbanismo / Prodam
-- Migration 004: Carga Inicial Real — Planilha OUCAE_ESTOQUE_abr_rv01.xlsx
-- PostgreSQL 15
-- Posição: 13/04/2026 | Gerado: {gerado_em}
--
-- ATENÇÃO: Execute APÓS 004a_revogar_seed_sintetico.sql.
-- ATENÇÃO: Revisar e aprovar antes de executar em produção.
--
-- Registros importados da ABA 2_CONTROLE_ESTOQUE:
--   - Propostas únicas inseridas (ON CONFLICT DO NOTHING)
--   - Certidões únicas inseridas (ON CONFLICT DO NOTHING)
--   - Títulos reais e movimentações de importação
-- Certidões da ABA 3_controle_certidoes acrescentadas ao final.
-- =============================================================================

BEGIN;

"""

FOOTER_004 = "\nCOMMIT;\n"


def escrever_migrations(sqls: list[str], output_dir: Path, gerado_em: str) -> None:
    path_004a = output_dir / "004a_revogar_seed_sintetico.sql"
    path_004 = output_dir / "004_carga_inicial_real.sql"

    path_004a.write_text(HEADER_004A, encoding="utf-8")
    print(f"  → {path_004a}")

    header = HEADER_004.format(gerado_em=gerado_em)
    body = "\n\n".join(sqls)
    path_004.write_text(header + body + FOOTER_004, encoding="utf-8")
    print(f"  → {path_004} ({len(sqls)} statements)")


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Gera migrations de carga inicial real do CEPAC.")
    parser.add_argument(
        "--xlsx",
        default="docs/OUCAE_ESTOQUE_abr_rv01.xlsx",
        help="Caminho para a planilha XLSX (default: docs/OUCAE_ESTOQUE_abr_rv01.xlsx)",
    )
    parser.add_argument(
        "--output-dir",
        default="migrations",
        help="Diretório de saída para os arquivos SQL (default: migrations/)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: Planilha não encontrada: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Lendo planilha: {xlsx_path}")
    sqls, totais = processar(xlsx_path)

    gerado_em = datetime.now().isoformat(timespec="seconds")
    print(f"\nEscrevendo migrations em {output_dir}/")
    escrever_migrations(sqls, output_dir, gerado_em)

    ok = imprimir_validacao(totais)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
